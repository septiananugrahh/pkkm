from openpyxl import load_workbook

# Load file Excel
wb = load_workbook("INSTRUMEN PKKM MASTER 1 tahunan.xlsx")
ws = wb.active  # ambil sheet pertama

output_sql = []
id_counter = 1
parent_unsur = None
parent_sub_unsur = None

def esc(text):
    if text is None:
        return ""
    return str(text).replace("'", "''")

# Mulai dari baris pertama supaya tidak hilang
for row in ws.iter_rows(min_row=1):
    col_a = row[0].value  # kolom A = Unsur
    col_c = row[2].value  # kolom C = Sub Unsur
    col_e = row[4].value  # kolom E = Indikator

    # Unsur
    if col_a:
        title = esc(col_a)
        output_sql.append(
            f"INSERT INTO nodes (id, parent_id, title, type, `order`, year_id, created_at, updated_at) "
            f"VALUES ({id_counter}, NULL, '{title}', 'unsur', 0, 1, NOW(), NOW());"
        )
        parent_unsur = id_counter
        id_counter += 1

    # Sub Unsur
    if col_c:
        title = esc(col_c)
        output_sql.append(
            f"INSERT INTO nodes (id, parent_id, title, type, `order`, year_id, created_at, updated_at) "
            f"VALUES ({id_counter}, {parent_unsur}, '{title}', 'sub_unsur', 0, 1, NOW(), NOW());"
        )
        parent_sub_unsur = id_counter
        id_counter += 1

    # Indikator
    if col_e:
        title = esc(col_e)
        output_sql.append(
            f"INSERT INTO nodes (id, parent_id, title, type, `order`, year_id, created_at, updated_at) "
            f"VALUES ({id_counter}, {parent_sub_unsur}, '{title}', 'indikator', 0, 1, NOW(), NOW());"
        )
        id_counter += 1

# Simpan ke file nodes.sql
with open("nodes.sql", "w", encoding="utf-8") as f:
    f.write("\n".join(output_sql))

print("✅ Selesai! File nodes.sql berhasil dibuat.")
